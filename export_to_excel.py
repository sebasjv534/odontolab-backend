"""
Script simplificado para generar reporte de actividades en Excel
Proyecto: OdontoLab Backend
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actividades"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colores calidad
    excelente = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bueno = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Encabezados
    headers = ["#", "Artefacto", "Actividad", "Nombre del Participante", 
               "Fecha de Inicio", "Fecha de Entrega Compromiso", 
               "Fecha de Entrega", "Porcentaje de Avance", "Calidad"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    # Datos
    data = [
        [1, "Arquitectura del Sistema", "Definición de arquitectura Clean Architecture y estructura de carpetas", 
         "Juan Sebastian Jimenez Villegas", "01/09/2025", "05/09/2025", "04/09/2025", "100%", "Excelente"],
        
        [2, "Configuración Base", "Setup inicial del proyecto FastAPI con dependencias (SQLAlchemy, python-jose, bcrypt)", 
         "Juan Sebastian Jimenez Villegas", "05/09/2025", "08/09/2025", "08/09/2025", "100%", "Bueno"],
        
        [3, "Core - Database", "Configuración de conexión async con PostgreSQL y pool de conexiones", 
         "Juan Sebastian Jimenez Villegas", "09/09/2025", "12/09/2025", "11/09/2025", "100%", "Excelente"],
        
        [4, "Core - Security", "Implementación de JWT authentication y hashing de contraseñas con bcrypt", 
         "Juan Sebastian Jimenez Villegas", "12/09/2025", "16/09/2025", "15/09/2025", "100%", "Excelente"],
        
        [5, "Modelos de Dominio", "Creación de modelos User, Patient, MedicalRecord, ContactRequest con SQLAlchemy", 
         "Juan Sebastian Jimenez Villegas", "16/09/2025", "22/09/2025", "21/09/2025", "100%", "Bueno"],
        
        [6, "Schemas Pydantic", "Definición de schemas de validación para todas las entidades (auth, users, patients)", 
         "Juan Sebastian Jimenez Villegas", "22/09/2025", "26/09/2025", "26/09/2025", "100%", "Bueno"],
        
        [7, "Repositorio - Users", "Implementación de UserRepository con operaciones CRUD asíncronas", 
         "Juan Sebastian Jimenez Villegas", "27/09/2025", "01/10/2025", "30/09/2025", "100%", "Excelente"],
        
        [8, "Repositorio - Patients", "Implementación de PatientRepository con búsqueda y filtros", 
         "Juan Sebastian Jimenez Villegas", "01/10/2025", "05/10/2025", "04/10/2025", "100%", "Bueno"],
        
        [9, "Servicio - Authentication", "Desarrollo de AuthService con login, registro y generación de tokens JWT", 
         "Juan Sebastian Jimenez Villegas", "05/10/2025", "10/10/2025", "09/10/2025", "100%", "Excelente"],
        
        [10, "Servicio - User Management", "Desarrollo de UserService con validaciones de seguridad (no self-delete, no last-admin)", 
         "Juan Sebastian Jimenez Villegas", "10/10/2025", "15/10/2025", "14/10/2025", "100%", "Excelente"],
        
        [11, "API Endpoints - Auth", "Implementación de endpoints /login, /register con OAuth2PasswordRequestForm", 
         "Juan Sebastian Jimenez Villegas", "15/10/2025", "19/10/2025", "18/10/2025", "100%", "Bueno"],
        
        [12, "API Endpoints - Users", "Implementación de CRUD completo de usuarios con serialización UUID/Enum", 
         "Juan Sebastian Jimenez Villegas", "19/10/2025", "24/10/2025", "23/10/2025", "100%", "Excelente"],
        
        [13, "API Endpoints - Patients", "Implementación de endpoints de pacientes con manejo de relaciones", 
         "Juan Sebastian Jimenez Villegas", "24/10/2025", "29/10/2025", "28/10/2025", "100%", "Bueno"],
        
        [14, "Deployment Setup", "Configuración de Render.com con Blueprint, PostgreSQL free tier y variables de entorno", 
         "Juan Sebastian Jimenez Villegas", "29/10/2025", "03/11/2025", "02/11/2025", "100%", "Excelente"],
        
        [15, "Database Initialization", "Creación de scripts init_db.py con retry logic y endpoint de emergency admin", 
         "Juan Sebastian Jimenez Villegas", "03/11/2025", "06/11/2025", "05/11/2025", "100%", "Excelente"],
        
        [16, "Testing - Postman", "Creación de colección Postman con 50+ requests y automatización de tokens", 
         "Juan Sebastian Jimenez Villegas", "06/11/2025", "10/11/2025", "09/11/2025", "100%", "Excelente"],
        
        [17, "Bug Fixes - Serialization", "Corrección de errores de serialización UUID y Enum en respuestas API", 
         "Juan Sebastian Jimenez Villegas", "10/11/2025", "13/11/2025", "12/11/2025", "100%", "Bueno"],
        
        [18, "Security Validations", "Implementación de validaciones para prevenir eliminación de admin y auto-eliminación", 
         "Juan Sebastian Jimenez Villegas", "13/11/2025", "17/11/2025", "16/11/2025", "100%", "Excelente"],
        
        [19, "Documentación Técnica", "Creación de documentación API (API_SUMMARY.md, FRONTEND_API_GUIDE.md, SECURITY_VALIDATIONS.md)", 
         "Juan Sebastian Jimenez Villegas", "17/11/2025", "22/11/2025", "21/11/2025", "100%", "Excelente"],
        
        [20, "Diseño de Base de Datos MVP", "Análisis y diseño de esquema completo con 15 tablas (diagramas ER en Mermaid y DBML)", 
         "Juan Sebastian Jimenez Villegas", "22/11/2025", "26/11/2025", "26/11/2025", "100%", "Excelente"]
    ]
    
    # Insertar datos
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            
            if col_idx in [1, 5, 6, 7, 8]:  # Columnas centradas
                cell.alignment = center
            else:
                cell.alignment = left
            
            # Color por calidad
            if col_idx == 9:  # Columna Calidad
                cell.font = Font(bold=True)
                if value == "Excelente":
                    cell.fill = excelente
                elif value == "Bueno":
                    cell.fill = bueno
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Altura de filas
    for row in range(2, 22):
        ws.row_dimensions[row].height = 40
    
    # Guardar
    filename = "OdontoLab_Actividades.xlsx"
    wb.save(filename)
    print(f"✅ Archivo generado: {filename}")
    print(f"📊 Contiene 20 actividades con formato profesional")

if __name__ == "__main__":
    create_excel()
